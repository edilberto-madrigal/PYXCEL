from openpyxl import Workbook as OpenpyxlWorkbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PySide6.QtCore import QObject, Signal
from PySide6.QtGui import QColor
import csv
from typing import Optional
import os

from ..models.spreadsheet import SpreadsheetModel, CellData, CellFormat
from ..models.workbook import Workbook


class FileManager(QObject):
    file_loaded = Signal(str)
    file_saved = Signal(str)
    error_occurred = Signal(str)

    def __init__(self):
        super().__init__()
        self._current_file_path: Optional[str] = None

    def open_file(self, file_path: str, workbook: Workbook) -> bool:
        try:
            ext = os.path.splitext(file_path)[1].lower()

            if ext == ".xlsx" or ext == ".xls":
                return self._open_excel(file_path, workbook)
            elif ext == ".csv":
                return self._open_csv(file_path, workbook)
            else:
                self.error_occurred.emit(f"Formato no soportado: {ext}")
                return False

        except Exception as e:
            self.error_occurred.emit(f"Error al abrir archivo: {str(e)}")
            return False

    def _open_excel(self, file_path: str, workbook: Workbook) -> bool:
        try:
            wb = load_workbook(file_path, data_only=True)

            workbook.clear()

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                max_row = ws.max_row
                max_col = ws.max_column

                sheet = workbook.add_sheet(sheet_name, max_row, max_col)

                for row in range(max_row):
                    for col in range(max_col):
                        cell = ws.cell(row + 1, col + 1)

                        if cell.value is not None:
                            cell_data = CellData()

                            if cell.data_type == "f":
                                cell_data.formula = cell.value
                                cell_data.is_calculated = False
                            else:
                                cell_data.value = cell.value
                                cell_data.is_calculated = True

                            self._apply_format_to_cell(cell, cell_data)

                            sheet._data[(row, col)] = cell_data

            wb.close()

            workbook.set_file_path(file_path)
            workbook.set_modified(False)
            self._current_file_path = file_path
            self.file_loaded.emit(file_path)

            return True

        except Exception as e:
            self.error_occurred.emit(f"Error al leer Excel: {str(e)}")
            return False

    def _apply_format_to_cell(self, openpyxl_cell, cell_data: CellData):
        if openpyxl_cell.font:
            cell_data.format.bold = openpyxl_cell.font.bold or False
            cell_data.format.italic = openpyxl_cell.font.italic or False
            cell_data.format.underline = bool(openpyxl_cell.font.underline)

            if openpyxl_cell.font.size:
                cell_data.format.font_size = openpyxl_cell.font.size

            if openpyxl_cell.font.color and openpyxl_cell.font.color.rgb:
                color = openpyxl_cell.font.color.rgb
                if len(color) == 6:
                    cell_data.format.font_color = QColor(
                        int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
                    )

        if openpyxl_cell.fill and openpyxl_cell.fill.fgColor:
            color = openpyxl_cell.fill.fgColor.rgb
            if color and len(color) == 6:
                cell_data.format.bg_color = QColor(
                    int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
                )

        if openpyxl_cell.alignment:
            h_align = openpyxl_cell.alignment.horizontal
            v_align = openpyxl_cell.alignment.vertical

            alignment = 0
            from PySide6.QtCore import Qt

            if h_align == "center":
                alignment |= Qt.AlignmentFlag.AlignHCenter
            elif h_align == "right":
                alignment |= Qt.AlignmentFlag.AlignRight
            else:
                alignment |= Qt.AlignmentFlag.AlignLeft

            if v_align == "center":
                alignment |= Qt.AlignmentFlag.AlignVCenter
            elif v_align == "bottom":
                alignment |= Qt.AlignmentFlag.AlignBottom
            else:
                alignment |= Qt.AlignmentFlag.AlignTop

            cell_data.format.alignment = alignment

    def _open_csv(self, file_path: str, workbook: Workbook) -> bool:
        try:
            workbook.clear()
            sheet = workbook.add_sheet("Hoja1")

            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader):
                    for col_idx, value in enumerate(row):
                        if value:
                            cell_data = CellData(value=value)
                            sheet._data[(row_idx, col_idx)] = cell_data

            workbook.set_file_path(file_path)
            workbook.set_modified(False)
            self._current_file_path = file_path
            self.file_loaded.emit(file_path)

            return True

        except Exception as e:
            self.error_occurred.emit(f"Error al leer CSV: {str(e)}")
            return False

    def save_file(self, file_path: str, workbook: Workbook) -> bool:
        try:
            ext = os.path.splitext(file_path)[1].lower()

            if ext == ".xlsx":
                return self._save_excel(file_path, workbook)
            elif ext == ".csv":
                return self._save_csv(file_path, workbook)
            else:
                self.error_occurred.emit(f"Formato no soportado: {ext}")
                return False

        except Exception as e:
            self.error_occurred.emit(f"Error al guardar: {str(e)}")
            return False

    def _save_excel(self, file_path: str, workbook: Workbook) -> bool:
        try:
            wb = OpenpyxlWorkbook()
            wb.remove(wb.active)

            for sheet_idx in range(workbook.sheet_count()):
                sheet = workbook.get_sheet(sheet_idx)
                sheet_name = workbook.get_sheet_name(sheet_idx)

                ws = wb.create_sheet(sheet_name)

                for (row, col), cell_data in sheet._data.items():
                    cell = ws.cell(row + 1, col + 1)

                    if cell_data.formula:
                        cell.value = f"={cell_data.formula}"
                    else:
                        cell.value = cell_data.value

                    self._apply_format_from_cell(cell_data, cell)

                for row in range(sheet._rows):
                    ws.row_dimensions[row + 1].height = 15

                for col in range(sheet._cols):
                    ws.column_dimensions[self._col_letter(col + 1)].width = 8.43

            wb.save(file_path)
            wb.close()

            workbook.set_file_path(file_path)
            workbook.set_modified(False)
            self._current_file_path = file_path
            self.file_saved.emit(file_path)

            return True

        except Exception as e:
            self.error_occurred.emit(f"Error al guardar Excel: {str(e)}")
            return False

    def _col_letter(self, col: int) -> str:
        result = ""
        while col > 0:
            result = chr(65 + ((col - 1) % 26)) + result
            col = (col - 1) // 26
        return result

    def _apply_format_from_cell(self, cell_data: CellData, openpyxl_cell):
        if cell_data.format.bold:
            openpyxl_cell.font = openpyxl_cell.font.copy(bold=True)
        if cell_data.format.italic:
            openpyxl_cell.font = openpyxl_cell.font.copy(italic=True)
        if cell_data.format.underline:
            openpyxl_cell.font = openpyxl_cell.font.copy(underline="single")

        if cell_data.format.font_size != 10:
            openpyxl_cell.font = openpyxl_cell.font.copy(
                size=cell_data.format.font_size
            )

        r, g, b, _ = cell_data.format.font_color.rgba()
        if (r, g, b) != (0, 0, 0):
            openpyxl_cell.font = openpyxl_cell.font.copy(color=f"{r:02x}{g:02x}{b:02x}")

        r, g, b, _ = cell_data.format.bg_color.rgba()
        if (r, g, b) != (255, 255, 255):
            openpyxl_cell.fill = PatternFill(
                start_color=f"{r:02x}{g:02x}{b:02x}",
                end_color=f"{r:02x}{g:02x}{b:02x}",
                fill_type="solid",
            )

    def _save_csv(self, file_path: str, workbook: Workbook) -> bool:
        try:
            sheet = workbook.current_sheet()
            if not sheet:
                return False

            with open(file_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)

                max_row = max([r for r, c in sheet._data.keys()], default=0) + 1
                max_col = max([c for r, c in sheet._data.keys()], default=0) + 1

                for row in range(max_row):
                    row_data = []
                    for col in range(max_col):
                        cell = sheet._data.get((row, col))
                        row_data.append(cell.value if cell else "")
                    writer.writerow(row_data)

            workbook.set_file_path(file_path)
            workbook.set_modified(False)
            self._current_file_path = file_path
            self.file_saved.emit(file_path)

            return True

        except Exception as e:
            self.error_occurred.emit(f"Error al guardar CSV: {str(e)}")
            return False

    def get_current_file_path(self) -> Optional[str]:
        return self._current_file_path
